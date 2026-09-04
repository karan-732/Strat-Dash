"""
Taking answers back from the client.

The consultant asks the questions a phase raised and comes back with something:
a filled spreadsheet, a call transcript, an email, or typed notes. This reads
that material against the open questions, closes the ones it genuinely answers,
records what the material added that nobody asked about, and rewrites the brain.
"""

from __future__ import annotations

import io
from typing import Any

from app.agents import answers as answer_agent
from app.agents import brain as brain_agent
from app.db import repo

TEXT_SUFFIXES = (".txt", ".md", ".markdown", ".vtt", ".srt", ".csv", ".tsv", ".json", ".log")
SHEET_SUFFIXES = (".xlsx", ".xlsm")


def extract_text(name: str, data: bytes, mime: str | None) -> str | None:
    """Pull readable text out of an upload, or None when it is not text-bearing."""
    lower = name.lower()

    if lower.endswith(SHEET_SUFFIXES):
        return _from_workbook(data)

    if lower.endswith(TEXT_SUFFIXES) or (mime or "").startswith("text/"):
        try:
            text = data.decode("utf-8", errors="replace")
        except Exception:  # noqa: BLE001
            return None
        return _clean_transcript(text)

    return None


def _clean_transcript(text: str) -> str:
    """Strip subtitle scaffolding so a VTT reads as speech."""
    import re

    text = re.sub(r"^WEBVTT.*$", "", text, flags=re.M | re.I)
    text = re.sub(r"^\s*\d+\s*$", "", text, flags=re.M)
    text = re.sub(r"^.*-->.*$", "", text, flags=re.M)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()[:24000]


def _from_workbook(data: bytes) -> str:
    """A filled answer sheet, flattened to `sheet | row` lines the agent can read."""
    try:
        from openpyxl import load_workbook

        book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return ""
    lines: list[str] = []
    for sheet in book.worksheets:
        lines.append(f"# sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c not in (None, "")]
            if cells:
                lines.append(" | ".join(cells))
            if len(lines) > 1200:
                break
    return "\n".join(lines)[:24000]


async def ingest_answers(
    *, engagement_id: str, material: str, material_name: str, source_ref: str
) -> dict[str, Any]:
    """Read returned material against the open questions and update the sprint."""
    engagement = await repo.get_engagement(engagement_id)
    if not engagement:
        raise ValueError("unknown engagement")

    open_questions = await repo.open_questions(engagement["id"])
    brain = await repo.current_brain(engagement["id"])

    result = await answer_agent.ingest(
        engagement=engagement,
        material=material,
        material_name=material_name,
        open_questions=open_questions,
        brain=brain,
    )
    await repo.log_run(
        engagement_id=engagement["id"],
        phase=None,
        agent="answers",
        status="ok",
        model=result["model"],
        usage=result["usage"],
    )

    """
    Everything the reader worked out, kept.

    It returns four things and only the answers used to be written down. The
    quotes, the half answers, the volunteered findings and the contradictions
    all went into the brain's prompt and then existed nowhere, so a refresh lost
    them and a partly-answered question looked untouched.
    """
    for hit in result["answered"]:
        await repo.answer_question(hit["id"], hit["answer"], source_ref, quote=str(hit.get("quote") or ""))
    for half in result["partial"]:
        await repo.mark_partial(
            half["id"],
            str(half.get("gotSoFar") or ""),
            str(half.get("stillMissing") or ""),
            source_ref,
        )
    await repo.save_answer_findings(
        engagement["id"],
        unprompted=result["unprompted"],
        contradictions=result["contradictions"],
        source_ref=source_ref,
    )

    landed = (
        f"MATERIAL RETURNED BY THE CLIENT — [{material_name}]\n\n"
        + "ANSWERS NOW SETTLED:\n"
        + ("\n".join(f'- {a["answer"]} (from: "{a["quote"][:180]}")' for a in result["answered"]) or "none")
        + "\n\nWHAT THE MATERIAL ADDED THAT NOBODY ASKED:\n"
        + ("\n".join(f"- {u['finding']}" for u in result["unprompted"]) or "none")
        + "\n\nWHAT IT CONTRADICTS:\n"
        + (
            "\n".join(
                f"- {c['finding']} (contradicts: {c.get('contradicts')})" for c in result["contradictions"]
            )
            or "none"
        )
    )

    revised = await brain_agent.revise(
        engagement=engagement, previous=brain, reason="answers_ingested", phase=None, landed=landed
    )
    version = await repo.save_brain(
        engagement_id=engagement["id"], reason="answers_ingested", phase=None, brain=revised
    )
    await repo.log_run(
        engagement_id=engagement["id"],
        phase=None,
        agent="brain",
        status="ok",
        model=revised["model"],
        usage=revised["usage"],
    )
    await repo.touch(engagement["id"])

    return {
        "answered": result["answered"],
        "partial": result["partial"],
        "unprompted": result["unprompted"],
        "contradictions": result["contradictions"],
        "stillOpen": len(open_questions) - len(result["answered"]),
        "brain": {
            "version": version,
            "narrative": revised.get("narrative", ""),
            "confidence": revised.get("confidence", 0),
        },
    }
